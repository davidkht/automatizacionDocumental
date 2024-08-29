import os
import sys
import math
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image  # Importa Image para añadir imágenes a los archivos Excel
from win32com import client
import pdfManager as pdf

PLANTILLA='plantillaLC.xlsx'
BASE_DE_DATOS='baseDeDatosParaRevisionDeListas.xlsx'

def xsl2pdf(file_location):
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False

    workbook=app.Workbooks.open(file_location)
    output = os.path.splitext(file_location)[0]

    # Configuración de impresión
    worksheet = workbook.ActiveSheet
    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = False  # Permitir que se extienda en múltiples páginas
    ##
    workbook.ActiveSheet.ExportAsFixedFormat(0,output)
    workbook.Close()
# Función para ajustar la altura de las celdas combinadas
def adjust_height(cell_range, text,sheet):
    # Asumir que cada columna tiene un ancho fijo estándar si no se especifica
    default_width = 8.43  # Este es el ancho estándar en Excel para columnas no ajustadas
    column_widths = [sheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width or default_width for cell in cell_range[0]]
    total_width = sum(column_widths)

    # Ajustar para mejor aproximación de caracteres por línea
    estimated_chars_per_line = total_width * 0.9  # Ajustar el factor según sea necesario

    # Calcular el número de líneas necesarias
    estimated_lines_needed = math.ceil(len(text) / estimated_chars_per_line)  # Estimar líneas necesarias basado en el ancho combinado
    
    # Añadir líneas adicionales por cada salto de línea en el texto
    explicit_lines = text.count('\n')

    # Total de líneas necesarias
    total_lines_needed = estimated_lines_needed + explicit_lines

    # Añadir margen adicional para asegurarse de que el texto es visible
    margin_factor = 1.2
    adjusted_lines_needed = total_lines_needed * margin_factor
    
    # Ajustar la altura de la fila
    cell_range[0][0].parent.row_dimensions[cell_range[0][0].row].height = adjusted_lines_needed * 15  # Ajustar la altura

def crear_lista_de_chequeo(pdseries,cliente, comercial, contrato, orden_c, ruta_de_guardado,directorio_script):

    ruta_plantilla=os.path.join(directorio_script,'docs',PLANTILLA)
    wb_LC = openpyxl.load_workbook(ruta_plantilla)  # Loads the Excel workbook.
    sheet= wb_LC.worksheets[0]

    ###########
    #ENCABEZADO
    ###########
    celda_cliente=sheet["E4:K4"]
    celda_cliente[0][0].value= cliente

    celda_comercial=sheet["E5:K5"]
    celda_comercial[0][0].value= comercial

    celda_ciudad=sheet["N4:Q4"]
    celda_ciudad[0][0].value= pdseries['CIUDAD']

    celda_fel=sheet["N5:Q5"]
    celda_fel[0][0].value= pdseries.iloc[7]

    celda_rev=sheet["N6:Q6"]
    celda_rev[0][0].value= pdseries.iloc[8]

    celda_contrato=sheet["E6:K6"]
    celda_contrato[0][0].value= contrato

    celda_orden=sheet["E7:K7"]
    celda_orden[0][0].value= orden_c

    celda_gestor=sheet["N7:Q7"]
    celda_gestor[0][0].value= pdseries.iloc[6]

    celda_ped=sheet["E8:K8"]
    celda_ped[0][0].value= "N/A"

    celda_caj=sheet["N8:Q8"]
    celda_caj[0][0].value= "N/A"

    celda_nombre=sheet["E10:K10"]
    celda_nombre[0][0].value= pdseries.iloc[2]

    celda_ref=sheet["N10:Q10"]
    celda_ref[0][0].value= pdseries.iloc[4]

    celda_marca=sheet["E11:K11"]
    celda_marca[0][0].value= pdseries.iloc[3]

    celda_serial=sheet["N11:Q11"]
    celda_serial[0][0].value= pdseries.iloc[5]


    ###################
    #CONEXION ELECTRICA
    ###################
    if pdseries.iloc[10]=='110V':
        sheet["H15"]='X'
    elif pdseries.iloc[10]=='220V':
        sheet["K15"]='X'

    sheet["M15"]=pdseries.iloc[11]

    celda_fase=sheet["P15:Q15"]
    celda_fase[0][0].value= pdseries.iloc[12]

    ###################
    #PARAMETROS
    ###################
    celdasDeX=['19','20','21','22','23','24','25','27','28','29','30','31','33']
    columnasEnBaseDeDatos=[13,15,17,19,21,23,25,27,29,31,33,35,37]
    for i in range(0,len(celdasDeX)):
        filaActual=celdasDeX[i]
        if pdseries.iloc[columnasEnBaseDeDatos[i]]=='SI':
            sheet['G'+filaActual]='X'

        elif pdseries.iloc[columnasEnBaseDeDatos[i]]=='NO':
            sheet['I'+filaActual]='X'

        else:
            sheet['K'+filaActual]='X'

        observaciones=sheet["M"+filaActual+":Q"+filaActual]
        observaciones[0][0].value= pdseries.iloc[columnasEnBaseDeDatos[i]+1]

    ###################
    #VARIABLES REV.
    ###################
    celda_variable1=sheet["C37:D37"]
    celda_variable1[0][0].value= pdseries.iloc[39]

    celda_variable1o=sheet["E37:K37"]
    celda_variable1o[0][0].value= pdseries.iloc[40]

    celda_variable2=sheet["C38:D38"]
    celda_variable2[0][0].value= pdseries.iloc[41]

    celda_variable2o=sheet["E38:K38"]
    celda_variable2o[0][0].value= pdseries.iloc[42]

    celda_variable3=sheet["M37:N37"]
    celda_variable3[0][0].value= pdseries.iloc[43]
    
    celda_variable3o=sheet["O37:Q37"]
    celda_variable3o[0][0].value= pdseries.iloc[44]

    celda_variable4=sheet["M38:N38"]
    celda_variable4[0][0].value= pdseries.iloc[45]
    
    celda_variable4o=sheet["O38:Q38"]
    celda_variable4o[0][0].value= pdseries.iloc[46]

    ######################
    #OBSERVACIONES Y FINAL
    ######################
    celda_observaciones=sheet["C41:Q44"]
    celda_observaciones[0][0].value= pdseries.iloc[47]

    celda_realizado=sheet["E52:M52"]
    celda_realizado[0][0].value= pdseries.iloc[48]

    celda_revisado=sheet["E53:M53"]
    celda_revisado[0][0].value= pdseries.iloc[49]

    consecutivo=str(pdseries.iloc[1])

    carpeta_de_almacenamiento= os.path.join(ruta_de_guardado,pdseries['CIUDAD'],consecutivo+" "+pdseries.iloc[2])
    sheet["J46"]=carpeta_de_almacenamiento
    sheet["J48"]=os.path.join(carpeta_de_almacenamiento,'REGISTRO AUDIOVISUAL')

    # Ajustar la altura de cada fila basándose en el contenido más largo que excede el ancho fijo
    # Ancho fijo de columna, por ejemplo, 20 caracteres
    fixed_column_width = 20
    for row in sheet.iter_rows(min_row=19):
        max_lines = 1  # Mínimo una línea por fila
        for cell in row:
            # Calcular cuántas líneas se necesitan para el texto en esta celda
            lines_needed = math.ceil(len(str(cell.value)) / fixed_column_width)
            if lines_needed > max_lines:
                max_lines = lines_needed
        # Ajustar la altura de la fila; asumir que cada línea necesita aproximadamente 15 puntos de altura
                if lines_needed==2:
                    sheet.row_dimensions[cell.row].height = max_lines * 15
                else:
                    sheet.row_dimensions[cell.row].height = max_lines * 9

    # adjust_height(celda_observaciones, str(pdseries.iloc[47]),sheet)
    # adjust_height(celda_variable1, str(pdseries.iloc[39]),sheet)
    # adjust_height(celda_variable1o, str(pdseries.iloc[40]),sheet)
    # adjust_height(celda_variable2, str(pdseries.iloc[41]),sheet)
    # adjust_height(celda_variable2o, str(pdseries.iloc[42]),sheet)
    # adjust_height(celda_variable3, str(pdseries.iloc[43]),sheet)
    # adjust_height(celda_variable3o, str(pdseries.iloc[44]),sheet)
    # adjust_height(celda_variable4, str(pdseries.iloc[45]),sheet)
    # adjust_height(celda_variable4o, str(pdseries.iloc[46]),sheet)


    img = Image(os.path.join(directorio_script, 'img', 'encabezadoLC.png'))
    sheet.add_image(img, 'C3')

    

    carpeta=os.path.join(directorio_script,'LC',consecutivo+" "+pdseries.iloc[2])
    os.makedirs(carpeta)
    os.makedirs(os.path.join(carpeta,'REGISTRO AUDIOVISUAL'))

    nombrearchivo=consecutivo+" "+pdseries.iloc[2]+'.xlsx'
    sheet.title=consecutivo
    archivo=os.path.join(carpeta,nombrearchivo)
    wb_LC.save(archivo)

    xsl2pdf(archivo)

def ejecutar_automatizacion_listasC(cliente,orden,contrato,ruta,comercial,directorio_script):
    ruta_basedatos=os.path.join(directorio_script,'..','basesDeDatos',BASE_DE_DATOS)
    #lee base de datos
    df = pd.read_excel(ruta_basedatos,keep_default_na=False)
    for indice, fila in df.iterrows():
        crear_lista_de_chequeo(fila,cliente,comercial,contrato,orden,ruta,directorio_script)

def buscar_pdf_y_registro(ruta_base):
    resultados = []
    
    # Recorre todos los subdirectorios de la ruta base
    for subdirectorio in os.listdir(ruta_base):
        ruta_subdirectorio = os.path.join(ruta_base, subdirectorio)                        
        if os.path.isdir(ruta_subdirectorio):
            # Buscar el archivo PDF en el subdirectorio
            archivo_pdf = None
            for archivo in os.listdir(ruta_subdirectorio):
                if archivo.endswith('.pdf'):
                    archivo_pdf = archivo
                    break
            
            if archivo_pdf:        
                resultados.append(( archivo_pdf, ruta_subdirectorio))
    
    return resultados

def unir_informe_con_fotos(directorio_script):

    registro_informes=os.path.join(directorio_script,'LC')

    rutas_subdirectorios=buscar_pdf_y_registro(registro_informes)
    
    for subdirectorio in rutas_subdirectorios:

        informe= os.path.join(subdirectorio[1],subdirectorio[0])
        registro_audiovisual = os.path.join(subdirectorio[1],'REGISTRO AUDIOVISUAL')
        archivo_de_imagenes  = os.path.join(registro_audiovisual,subdirectorio[0])
        pdf.insert_images_to_pdf(registro_audiovisual,archivo_de_imagenes)
        pdf.merge_pdfs([informe,archivo_de_imagenes] , informe)